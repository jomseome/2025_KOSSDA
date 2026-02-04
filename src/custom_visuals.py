from __future__ import annotations

from functools import lru_cache
import io
import re
from typing import Dict

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from openpyxl import load_workbook
from PIL import Image

from .workspace_data import DATA_DIR

DATA_PATH = DATA_DIR / "excel_data" / "sample_story_points.xlsx"
EDUCATION_CARE_PATH = (
    DATA_DIR
    / "excel_data"
    / "(0204)교육돌봄 데이터_v1.0.xlsx"
)
POLITICS_CIVIC_PATH = (
    DATA_DIR
    / "excel_data"
    / "(0204)정치사회 데이터_v1.0.xlsx"
)

EDUCATION_CARE_SHEETS = {
    1: "그림1",
    2: "그림2",
    3: "그림3",
    4: "그림4",
    5: "그림5",
    6: "그림6",
    7: "그림7",
    8: "그림8",
    9: "그림9",
    10: "그림10",
    11: "그림11",
    12: "그림12",
    13: "그림13",
    14: "그림14",
    15: "그림 15",
}

EDUCATION_CARE_TITLES = {
    1: "그림 1. 학교급별 학령인구 수, 2015-2035",
    2: "그림 2. 한국과 OECD 국가의 초·중·고 교사 1인당 학생 수, 2021-2023",
    3: "그림 3. 한국과 OECD 국가의 초·중학교 학급당 학생 수, 2019-2023",
    4: "그림 4. 한국과 OECD 국가의 교육단계별 GDP 대비 공교육비 비율, 2022",
    5: "그림 5. 한국과 OECD 국가의 국제 학업성취도(PISA) 결과, 2015-2022",
    6: "그림 6. 한국과 OECD 국가의 수학 점수 분산 비율, 2012-2022",
    7: "그림 7. 학교급별 학생 1인당 월평균 사교육비, 2010-2024",
    8: "그림 8. OECD 국가의 사교육 참여 수준, 2022",
    9: "그림 9. 초·중·고 학교에 대한 평가, 2024",
    10: "그림 10. OECD 주요국 학생의 학업 관련 불안감, 2015",
    11: "그림 11. OECD 주요국 학부모의 학교교육 만족도, 2009-2022",
    12: "그림 12. 교육수준별 전공-직업 일치도, 2000-2024",
    13: "그림 13. 보육 및 유아교육 기관 만족도, 2015-2024",
    14: "그림 14. 주요국 학생의 학교 밖 신체활동 미참여율, 2015",
    15: "그림 15. OECD 주요국의 GDP 대비 보육·유아교육 공공지출 비율",
}

POLITICS_CIVIC_SHEETS = {idx: f"그림{idx}" for idx in range(1, 15)}

POLITICS_CIVIC_TITLES = {
    1: "그림 1. 일반 신뢰, 2003-2025",
    2: "그림 2. 이타주의와 이기주의에 대한 인식, 2003-2025",
    3: "그림 3. 공정성에 대한 인식, 2003-2025",
    4: "그림 4. 중앙정부에 대한 신뢰, 2003-2025",
    5: "그림 5. 대통령실에 대한 신뢰, 2003-2025",
    6: "그림 6. 국회에 대한 신뢰, 2003-2025",
    7: "그림 7. 대법원에 대한 신뢰, 2003-2025",
    8: "그림 8. 이념 성향별 정치 효능감, 2013-2024",
    9: "그림 9. 선거 투표율, 1987-2025",
    10: "그림 10. 비선거적 정치 참여율, 2013-2024",
    11: "그림 11. 무당파의 이념 성향, 2003-2025",
    12: "그림 12. 진보 정당 지지자의 이념 성향, 2003-2025",
    13: "그림 13. 보수 정당 지지자의 이념 성향, 2003-2025",
    14: "그림 14. 주요국의 정치 양극화 수준, 2000-2025",
}


@lru_cache(maxsize=4)
def _load_dataset() -> Dict[str, pd.DataFrame]:
    if not DATA_PATH.exists():
        raise FileNotFoundError(f"샘플 데이터 파일을 찾을 수 없습니다: {DATA_PATH}")
    excel = pd.ExcelFile(DATA_PATH)
    return {name: excel.parse(name) for name in excel.sheet_names}


@lru_cache(maxsize=1)
def _education_care_excel() -> pd.ExcelFile:
    if not EDUCATION_CARE_PATH.exists():
        raise FileNotFoundError(f"교육·돌봄 데이터 파일을 찾을 수 없습니다: {EDUCATION_CARE_PATH}")
    return pd.ExcelFile(EDUCATION_CARE_PATH)


@lru_cache(maxsize=32)
def _education_care_sheet(sheet: str) -> pd.DataFrame:
    excel = _education_care_excel()
    if sheet not in excel.sheet_names:
        raise KeyError(f"시트 `{sheet}`을(를) 찾을 수 없습니다.")
    return excel.parse(sheet)


@lru_cache(maxsize=1)
def _politics_civic_excel() -> pd.ExcelFile:
    if not POLITICS_CIVIC_PATH.exists():
        raise FileNotFoundError(f"정치·시민사회 데이터 파일을 찾을 수 없습니다: {POLITICS_CIVIC_PATH}")
    return pd.ExcelFile(POLITICS_CIVIC_PATH)


@lru_cache(maxsize=32)
def _politics_civic_sheet(sheet: str) -> pd.DataFrame:
    excel = _politics_civic_excel()
    if sheet not in excel.sheet_names:
        raise KeyError(f"시트 `{sheet}`을(를) 찾을 수 없습니다.")
    return excel.parse(sheet)


@lru_cache(maxsize=1)
def _education_care_sheet_image(sheet: str) -> Image.Image | None:
    if not EDUCATION_CARE_PATH.exists():
        return None
    wb = load_workbook(EDUCATION_CARE_PATH, data_only=True)
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    images = getattr(ws, "_images", [])
    if not images:
        return None
    img = images[0]
    data_attr = getattr(img, "_data", None)
    image_data = data_attr() if callable(data_attr) else data_attr
    if not image_data:
        return None
    return Image.open(io.BytesIO(image_data)).convert("RGBA")


def _parse_year(value) -> float:
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return int(value)
    match = re.search(r"(\d{4})", str(value).strip())
    return int(match.group(1)) if match else np.nan


def _year_columns(columns) -> list:
    return [col for col in columns if not pd.isna(_parse_year(col))]


def _clean_unnamed_first_col(df: pd.DataFrame, column_name: str = "category") -> pd.DataFrame:
    df = df.copy()
    if "Unnamed: 0" in df.columns:
        return df.rename(columns={"Unnamed: 0": column_name})
    return df.rename(columns={df.columns[0]: column_name})


def _to_long_wide_years(df_wide: pd.DataFrame, id_col: str, value_name: str = "value") -> pd.DataFrame:
    df = df_wide.copy()
    year_cols = _year_columns(df.columns)
    long_df = df.melt(id_vars=[id_col], value_vars=year_cols, var_name="year", value_name=value_name)
    long_df["year"] = long_df["year"].apply(_parse_year)
    long_df = long_df.dropna(subset=["year"])
    long_df["year"] = long_df["year"].astype(int)
    return long_df


def _apply_common_layout(fig: go.Figure, title: str) -> go.Figure:
    fig.update_layout(
        title=title,
        hovermode="closest",
        legend_title_text="",
        margin=dict(l=40, r=20, t=70, b=40),
    )
    return fig


def _ensure_dataset(sheet: str) -> pd.DataFrame:
    data_map = _load_dataset()
    if sheet not in data_map:
        raise KeyError(f"시트 `{sheet}`을(를) 찾을 수 없습니다.")
    return data_map[sheet]


def covid_section2_chart(story_slug: str, slot_id: str) -> go.Figure:
    df = _ensure_dataset("crime_trend")
    fig = px.line(
        df,
        x="year",
        y=["전국", "서울"],
        markers=True,
        color_discrete_map={
            "전국": "#1565C0",  # deep blue
            "서울": "#0D47A1",  # slightly darker blue for contrast
        },
        title="세계 주요 도시 추세와 비교: 전국 vs 서울",
    )
    fig.update_layout(
        xaxis_title="연도",
        yaxis_title="1인당 지표",
        font={"family": "Noto Sans KR, Noto Sans, sans-serif"},
    )
    return fig


def covid_section3_chart(story_slug: str, slot_id: str) -> go.Figure:
    df = _ensure_dataset("welfare_overview")
    fig = px.area(
        df,
        x="year",
        y="사회지출(GDP%)",
        title="한국 범죄유형별 변화에 영향을 준 사회지출 흐름",
    )
    fig.update_traces(
        line={"color": "#0D47A1", "width": 3},
        fillcolor="rgba(21, 101, 192, 0.25)",  # light blue fill
    )
    fig.update_layout(
        xaxis_title="연도",
        yaxis_title="사회지출(GDP%)",
        font={"family": "Noto Sans KR, Noto Sans, sans-serif"},
    )
    return fig


def covid_interactive_panel(story_slug: str) -> None:
    df_crime = _ensure_dataset("crime_trend")
    df_welfare = _ensure_dataset("welfare_overview")

    st.markdown("#### 데이터 탐색")
    st.caption("연도 범위를 조정하거나 항목을 골라서 살펴보세요.")

    with st.container():
        col1, col2 = st.columns(2)
        with col1:
            metric = st.selectbox("범죄 지표", options=["전국", "서울", "부산"], key="covid_metric")
        with col2:
            year_range = st.slider(
                "연도 범위",
                min_value=int(df_crime["year"].min()),
                max_value=int(df_crime["year"].max()),
                value=(2014, 2024),
            )

    mask = df_crime["year"].between(*year_range)
    fig = px.line(
        df_crime[mask],
        x="year",
        y=metric,
        markers=True,
        title=f"{metric} 범죄 지표 추세",
    )
    fig.update_layout(xaxis_title="연도", yaxis_title="지표 값")
    st.plotly_chart(fig, use_container_width=True, key="covid_interactive_crime")

    st.markdown("##### 복지 지표와 비교")
    fig_welfare = px.area(
        df_welfare[mask],
        x="year",
        y=["사회지출(GDP%)", "빈곤율(%)"],
        title="사회지출과 빈곤율",
    )
    fig_welfare.update_layout(xaxis_title="연도")
    st.plotly_chart(fig_welfare, use_container_width=True, key="covid_interactive_welfare")


def education_care_fig01(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[1]).rename(columns={"Unnamed: 0": "구분"})
    year_cols = _year_columns(df.columns)
    long = df.melt(id_vars=["구분"], value_vars=year_cols, var_name="연도", value_name="값")

    df_bar = long[long["구분"].isin(["유치원", "초등학교", "중학교", "고등학교"])].copy()
    df_total = long[long["구분"] == "전체"].copy()

    fig = go.Figure()
    for cat in ["유치원", "초등학교", "중학교", "고등학교"]:
        tmp = df_bar[df_bar["구분"] == cat]
        fig.add_trace(go.Bar(x=tmp["연도"], y=tmp["값"], name=cat))

    fig.add_trace(
        go.Scatter(
            x=df_total["연도"],
            y=df_total["값"],
            mode="lines+markers",
            name="전체(라인)",
            yaxis="y2",
        )
    )

    fig.update_layout(
        barmode="stack",
        xaxis_title="연도",
        yaxis_title="학령인구",
        yaxis2=dict(title="전체", overlaying="y", side="right", showgrid=False),
    )
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[1])


def education_care_fig02(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[2]).copy()
    df["기준연도"] = df["기준연도"].ffill()
    long = df.melt(
        id_vars=["기준연도", "구분"],
        value_vars=["초등학교", "중학교", "고등학교"],
        var_name="학교급",
        value_name="교사1인당학생수",
    )

    fig = px.bar(
        long,
        x="학교급",
        y="교사1인당학생수",
        color="구분",
        barmode="group",
        facet_col="기준연도",
    )
    fig.update_layout(height=520)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[2])


def education_care_fig03(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[3]).copy()
    df["기준연도"] = df["기준연도"].ffill()
    long = df.melt(
        id_vars=["기준연도", "구분"],
        value_vars=["초등학교", "중학교"],
        var_name="학교급",
        value_name="학급당학생수",
    )

    fig = px.bar(
        long,
        x="학교급",
        y="학급당학생수",
        color="구분",
        barmode="group",
        facet_col="기준연도",
    )
    fig.update_layout(height=520)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[3])


def education_care_fig04(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[4])

    df2 = df.iloc[1:].copy()
    df2 = df2.rename(columns={"구분": "국가"})
    df2.columns = ["국가"] + df2.columns[1:].tolist()

    stage_map = [
        ("초·중등교육", "정부"),
        ("초·중등교육", "민간"),
        ("초·중등교육", "합계"),
        ("고등교육", "정부"),
        ("고등교육", "민간"),
        ("고등교육", "합계"),
        ("초등~고등교육", "정부"),
        ("초등~고등교육", "민간"),
        ("초등~고등교육", "합계"),
    ]

    value_cols = df2.columns[1:]
    if len(value_cols) != 9:
        raise ValueError(f"그림4: 예상 컬럼(9개)과 다릅니다. 현재={len(value_cols)}")

    records = []
    for (stage, fund), col in zip(stage_map, value_cols):
        tmp = df2[["국가", col]].copy()
        tmp["교육단계"] = stage
        tmp["재원"] = fund
        tmp = tmp.rename(columns={col: "GDP대비비율(%)"})
        records.append(tmp)
    long = pd.concat(records, ignore_index=True)

    fig = px.bar(
        long,
        x="교육단계",
        y="GDP대비비율(%)",
        color="재원",
        barmode="group",
        facet_col="국가",
    )
    fig.update_layout(height=520)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[4])


def education_care_fig05(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[5])
    long = df.melt(
        id_vars=["연도", "영역"],
        value_vars=["한국", "OECD 평균"],
        var_name="구분",
        value_name="점수",
    )
    fig = px.bar(
        long,
        x="영역",
        y="점수",
        color="구분",
        barmode="group",
        facet_col="연도",
    )
    fig.update_layout(height=520)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[5])


def education_care_fig06(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[6])
    df2 = df.copy()
    df2 = df2.rename(columns={"Unnamed: 0": "지표"})

    colmap = {
        2012: ("2012", "한국"),
        "Unnamed: 2": ("2012", "OECD 평균"),
        2022: ("2022", "한국"),
        "Unnamed: 4": ("2022", "OECD 평균"),
    }

    records = []
    for col, (year, group) in colmap.items():
        tmp = df2[["지표", col]].copy()
        tmp["연도"] = year
        tmp["구분"] = group
        tmp = tmp.rename(columns={col: "값"})
        records.append(tmp)
    long = pd.concat(records, ignore_index=True)

    fig = px.bar(
        long,
        x="지표",
        y="값",
        color="구분",
        barmode="group",
        facet_col="연도",
    )
    fig.update_layout(height=520)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[6])


def education_care_fig07(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[7]).copy()
    df = df.rename(columns={"Unnamed: 0": "구분"})
    year_cols = _year_columns(df.columns)
    long = df.melt(id_vars=["구분"], value_vars=year_cols, var_name="연도", value_name="사교육비")

    fig = px.line(
        long,
        x="연도",
        y="사교육비",
        color="구분",
        markers=True,
    )
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[7])


def education_care_fig08(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[8]).copy()
    df = df.sort_values("사교육 참여수준", ascending=True)

    fig = px.bar(
        df,
        x="사교육 참여수준",
        y="국가",
        orientation="h",
    )
    fig.update_layout(height=900)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[8])


def education_care_fig09(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[9]).copy()
    df["평가대상"] = df["평가대상"].ffill()

    dist_cols = ["매우 잘하고 있다", "잘하고 있다", "보통이다", "못하고 있다", "전혀 못하고 있다"]

    dist_long = df.melt(
        id_vars=["평가대상", "구분", "2023년 평균점수", "2024년 평균점수"],
        value_vars=dist_cols,
        var_name="응답",
        value_name="비율",
    )

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    dist_long["대상"] = dist_long["평가대상"].astype(str) + " / " + dist_long["구분"].astype(str)

    for resp in dist_cols:
        tmp = dist_long[dist_long["응답"] == resp]
        fig.add_trace(go.Bar(x=tmp["대상"], y=tmp["비율"], name=resp), secondary_y=False)

    base = df.copy()
    base["대상"] = base["평가대상"].astype(str) + " / " + base["구분"].astype(str)

    fig.add_trace(
        go.Scatter(x=base["대상"], y=base["2023년 평균점수"], mode="lines+markers", name="2023 평균점수"),
        secondary_y=True,
    )
    fig.add_trace(
        go.Scatter(x=base["대상"], y=base["2024년 평균점수"], mode="lines+markers", name="2024 평균점수"),
        secondary_y=True,
    )

    fig.update_layout(barmode="stack", xaxis_title="평가대상 / 구분", legend_title="항목")
    fig.update_yaxes(title_text="응답 비율(%)", secondary_y=False)
    fig.update_yaxes(title_text="평균점수", secondary_y=True, showgrid=False)

    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[9])


def education_care_fig10(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[10]).copy()
    bar_cols = ["시험을 잘 준비했더라도 불안하다", "공부를 할 때 매우 긴장된다"]
    line_col = "학업 관련 불안감 지수"

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    for col in bar_cols:
        fig.add_trace(go.Bar(x=df["국가"], y=df[col], name=col), secondary_y=False)

    fig.add_trace(
        go.Scatter(x=df["국가"], y=df[line_col], mode="lines+markers", name=line_col),
        secondary_y=True,
    )

    fig.update_layout(barmode="group", xaxis_title="국가")
    fig.update_yaxes(title_text="비율(%)", secondary_y=False)
    fig.update_yaxes(title_text="불안감 지수", secondary_y=True, showgrid=False)

    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[10])


def education_care_fig11(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[11]).copy()
    df = df.rename(columns={"Unnamed: 0": "국가"})
    year_cols = _year_columns(df.columns)
    long = df.melt(id_vars=["국가"], value_vars=year_cols, var_name="연도", value_name="만족도")
    long["연도"] = long["연도"].apply(_parse_year)
    long = long.dropna(subset=["연도"])

    fig = px.line(
        long,
        x="연도",
        y="만족도",
        color="국가",
        markers=True,
    )
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[11])


def education_care_fig12(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[12]).copy()
    df = df.rename(columns={"Unnamed: 0": "교육수준"})
    year_cols = _year_columns(df.columns)
    long = df.melt(id_vars=["교육수준"], value_vars=year_cols, var_name="연도", value_name="일치도")
    long["연도"] = long["연도"].apply(_parse_year)
    long = long.dropna(subset=["연도"])

    fig = px.line(
        long,
        x="연도",
        y="일치도",
        color="교육수준",
        markers=True,
    )
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[12])


def education_care_fig13(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[13]).copy()
    long = df.melt(
        id_vars=["구분"],
        value_vars=["전체", "어린이집", "유치원"],
        var_name="기관",
        value_name="만족도",
    )
    fig = px.line(
        long,
        x="구분",
        y="만족도",
        color="기관",
        markers=True,
    )
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[13])


def education_care_fig14(story_slug: str, slot_id: str) -> go.Figure:
    df = _education_care_sheet(EDUCATION_CARE_SHEETS[14]).copy()
    df = df.rename(columns={"Unnamed: 0": "국가"})
    df = df.sort_values("Girls", ascending=True)
    long = df.melt(id_vars=["국가"], value_vars=["Boys", "Girls"], var_name="성별", value_name="미참여율")

    fig = px.bar(
        long,
        x="미참여율",
        y="국가",
        color="성별",
        orientation="h",
        barmode="group",
    )
    fig.update_layout(height=900)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[14])


def education_care_fig15(story_slug: str, slot_id: str) -> go.Figure:
    img = _education_care_sheet_image(EDUCATION_CARE_SHEETS[15])
    if img is None:
        fig = go.Figure()
        fig.add_annotation(
            text="이미지를 찾을 수 없습니다.",
            x=0.5,
            y=0.5,
            showarrow=False,
            xref="paper",
            yref="paper",
        )
        return _apply_common_layout(fig, EDUCATION_CARE_TITLES[15])

    fig = go.Figure(go.Image(z=np.array(img)))
    fig.update_xaxes(showticklabels=False, showgrid=False, zeroline=False)
    fig.update_yaxes(showticklabels=False, showgrid=False, zeroline=False)
    return _apply_common_layout(fig, EDUCATION_CARE_TITLES[15])


def politics_civic_fig01(story_slug: str, slot_id: str) -> go.Figure:
    df = _politics_civic_sheet(POLITICS_CIVIC_SHEETS[1])
    df = df.dropna(how="all")
    if list(df.columns) == [0, 1]:
        df.columns = ["year", "score"]
    else:
        df = df.rename(columns={df.columns[0]: "year", df.columns[1]: "score"})
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["year"])
    df["year"] = df["year"].astype(int)

    fig = px.line(df, x="year", y="score", markers=True)
    fig.update_layout(xaxis_title="연도", yaxis_title="점수")
    return _apply_common_layout(fig, POLITICS_CIVIC_TITLES[1])


def _politics_stacked(sheet: str, title: str) -> go.Figure:
    df = _politics_civic_sheet(sheet)
    df = _clean_unnamed_first_col(df)
    long = _to_long_wide_years(df, id_col="category", value_name="percent")

    fig = px.bar(long, x="year", y="percent", color="category", barmode="stack")
    fig.update_layout(xaxis_title="연도", yaxis_title="비율(%)")
    fig.update_yaxes(range=[0, 100])
    return _apply_common_layout(fig, title)


def politics_civic_fig02(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_stacked(POLITICS_CIVIC_SHEETS[2], POLITICS_CIVIC_TITLES[2])


def politics_civic_fig03(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_stacked(POLITICS_CIVIC_SHEETS[3], POLITICS_CIVIC_TITLES[3])


def politics_civic_fig04(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_stacked(POLITICS_CIVIC_SHEETS[4], POLITICS_CIVIC_TITLES[4])


def politics_civic_fig05(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_stacked(POLITICS_CIVIC_SHEETS[5], POLITICS_CIVIC_TITLES[5])


def politics_civic_fig06(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_stacked(POLITICS_CIVIC_SHEETS[6], POLITICS_CIVIC_TITLES[6])


def politics_civic_fig07(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_stacked(POLITICS_CIVIC_SHEETS[7], POLITICS_CIVIC_TITLES[7])


def politics_civic_fig08(story_slug: str, slot_id: str) -> go.Figure:
    df = _politics_civic_sheet(POLITICS_CIVIC_SHEETS[8])
    df = _clean_unnamed_first_col(df)
    long = _to_long_wide_years(df, id_col="category", value_name="score")

    fig = px.line(long, x="year", y="score", color="category", markers=True)
    fig.update_layout(xaxis_title="연도", yaxis_title="점수")
    return _apply_common_layout(fig, POLITICS_CIVIC_TITLES[8])


def politics_civic_fig09(story_slug: str, slot_id: str) -> go.Figure:
    df = _politics_civic_sheet(POLITICS_CIVIC_SHEETS[9]).copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = _clean_unnamed_first_col(df, column_name="year")
    long = df.melt(id_vars=["year"], var_name="election_type", value_name="turnout")
    long["year"] = pd.to_numeric(long["year"], errors="coerce")
    long = long.dropna(subset=["year"])
    long["year"] = long["year"].astype(int)

    fig = px.scatter(long, x="year", y="turnout", color="election_type")
    fig.update_traces(mode="lines+markers")
    fig.update_layout(xaxis_title="연도", yaxis_title="투표율(%)")
    fig.update_yaxes(range=[0, 100])
    return _apply_common_layout(fig, POLITICS_CIVIC_TITLES[9])


def _parse_politics_fig10_blocks() -> pd.DataFrame:
    raw = _politics_civic_sheet(POLITICS_CIVIC_SHEETS[10])
    raw = raw.replace({np.nan: None})
    records = []

    def collect_block(activity, year_col, rate_col, start_row):
        r = start_row
        while r < len(raw):
            y = raw.iloc[r, year_col]
            v = raw.iloc[r, rate_col]
            if (raw.iloc[r, 0] is not None) or (raw.iloc[r, 4] is not None):
                if r != start_row:
                    break
            if y is None and v is None:
                r += 1
                continue
            try:
                year = int(float(y))
                rate = float(v)
                records.append({"activity": activity, "year": year, "rate": rate})
            except Exception:
                pass
            r += 1
        return r

    r = 0
    while r < len(raw):
        left_name = raw.iloc[r, 0]
        right_name = raw.iloc[r, 4]
        if left_name is not None and isinstance(left_name, str) and left_name.strip():
            r = collect_block(left_name.strip(), 1, 2, r + 1)
            continue
        if right_name is not None and isinstance(right_name, str) and right_name.strip():
            r = collect_block(right_name.strip(), 5, 6, r + 1)
            continue
        r += 1

    return pd.DataFrame.from_records(records)


def politics_civic_fig10(story_slug: str, slot_id: str) -> go.Figure:
    long = _parse_politics_fig10_blocks()
    long = long.sort_values(["activity", "year"]).reset_index(drop=True)

    activities = long["activity"].unique().tolist()
    default_act = activities[0] if activities else None

    fig = go.Figure()
    for act in activities:
        d = long[long["activity"] == act]
        fig.add_trace(
            go.Scatter(
                x=d["year"],
                y=d["rate"],
                mode="lines+markers",
                name=act,
                visible=(act == default_act),
            )
        )

    buttons = []
    for i, act in enumerate(activities):
        vis = [False] * len(activities)
        vis[i] = True
        buttons.append(
            dict(
                label=act,
                method="update",
                args=[{"visible": vis}, {"title": f"{POLITICS_CIVIC_TITLES[10]} - {act}"}],
            )
        )

    fig.update_layout(
        title=f"{POLITICS_CIVIC_TITLES[10]} - {default_act}" if default_act else POLITICS_CIVIC_TITLES[10],
        xaxis_title="연도",
        yaxis_title="참여율(%)",
        updatemenus=[dict(buttons=buttons, direction="down", x=1.02, y=1.0, xanchor="left")],
    )
    if not long.empty:
        fig.update_yaxes(range=[0, max(10, float(long["rate"].max()))])
    return _apply_common_layout(fig, POLITICS_CIVIC_TITLES[10])


def _politics_ideology(sheet: str, title: str) -> go.Figure:
    raw = _politics_civic_sheet(sheet).dropna(how="all")
    header = raw.iloc[1].tolist()
    cols = ["year"] + [str(h).strip() for h in header[1:]]
    df = raw.iloc[2:].copy()
    df.columns = cols
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["year"])
    df["year"] = df["year"].astype(int)
    long = df.melt(id_vars=["year"], var_name="ideology", value_name="percent")
    long["percent"] = pd.to_numeric(long["percent"], errors="coerce")

    fig = px.line(long, x="year", y="percent", color="ideology", markers=True)
    fig.update_layout(xaxis_title="연도", yaxis_title="비율(%)")
    fig.update_yaxes(range=[0, 100])
    return _apply_common_layout(fig, title)


def politics_civic_fig11(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_ideology(POLITICS_CIVIC_SHEETS[11], POLITICS_CIVIC_TITLES[11])


def politics_civic_fig12(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_ideology(POLITICS_CIVIC_SHEETS[12], POLITICS_CIVIC_TITLES[12])


def politics_civic_fig13(story_slug: str, slot_id: str) -> go.Figure:
    return _politics_ideology(POLITICS_CIVIC_SHEETS[13], POLITICS_CIVIC_TITLES[13])


def politics_civic_fig14(story_slug: str, slot_id: str) -> go.Figure:
    df = _politics_civic_sheet(POLITICS_CIVIC_SHEETS[14])
    df = _clean_unnamed_first_col(df, column_name="year")
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.dropna(subset=["year"])
    df["year"] = df["year"].astype(int)
    long = df.melt(id_vars=["year"], var_name="country", value_name="polarization")
    long["polarization"] = pd.to_numeric(long["polarization"], errors="coerce")

    fig = px.scatter(long, x="year", y="polarization", color="country")
    fig.update_traces(mode="lines+markers")
    fig.update_layout(xaxis_title="연도", yaxis_title="양극화 지표")
    return _apply_common_layout(fig, POLITICS_CIVIC_TITLES[14])


VISUAL_RENDERERS = {
    "covid_section2_chart": covid_section2_chart,
    "covid_section3_chart": covid_section3_chart,
    "education_care_fig01": education_care_fig01,
    "education_care_fig02": education_care_fig02,
    "education_care_fig03": education_care_fig03,
    "education_care_fig04": education_care_fig04,
    "education_care_fig05": education_care_fig05,
    "education_care_fig06": education_care_fig06,
    "education_care_fig07": education_care_fig07,
    "education_care_fig08": education_care_fig08,
    "education_care_fig09": education_care_fig09,
    "education_care_fig10": education_care_fig10,
    "education_care_fig11": education_care_fig11,
    "education_care_fig12": education_care_fig12,
    "education_care_fig13": education_care_fig13,
    "education_care_fig14": education_care_fig14,
    "education_care_fig15": education_care_fig15,
    "politics_civic_fig01": politics_civic_fig01,
    "politics_civic_fig02": politics_civic_fig02,
    "politics_civic_fig03": politics_civic_fig03,
    "politics_civic_fig04": politics_civic_fig04,
    "politics_civic_fig05": politics_civic_fig05,
    "politics_civic_fig06": politics_civic_fig06,
    "politics_civic_fig07": politics_civic_fig07,
    "politics_civic_fig08": politics_civic_fig08,
    "politics_civic_fig09": politics_civic_fig09,
    "politics_civic_fig10": politics_civic_fig10,
    "politics_civic_fig11": politics_civic_fig11,
    "politics_civic_fig12": politics_civic_fig12,
    "politics_civic_fig13": politics_civic_fig13,
    "politics_civic_fig14": politics_civic_fig14,
}


INTERACTIVE_PANELS = {
    "covid19": covid_interactive_panel,
}
